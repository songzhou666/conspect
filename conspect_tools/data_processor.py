"""
vision_tools 数据处理模块
提供Excel解析、数据清洗、维度识别、指标计算等功能

核心设计原则：
  - 不硬编码任何业务数据
  - 不假设列名或数据类型
  - 所有方法都可通过参数自定义行为
"""
import pandas as pd
import openpyxl
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple, Union


class DataProcessor:
    """Excel数据处理核心类"""

    def __init__(self):
        self.dataframes: Dict[str, pd.DataFrame] = {}
        self.merged_df: Optional[pd.DataFrame] = None

    def load_excel(self, file_paths: List[str]) -> Dict[str, pd.DataFrame]:
        """
        加载多个Excel文件，读取所有Sheet的所有数据。

        使用 openpyxl 引擎读取，保留原始列名和行顺序。
        不做任何清洗或转换——仅读取原始数据。
        数值列可能被 pandas 推断为 int64/float64，日期列推断为 datetime64，
        文本列保持为 object 类型。

        参数:
            file_paths: Excel文件路径列表
        返回:
            {文件名_Sheet名: DataFrame} 字典，包含所有文件的所有Sheet
        """
        result = {}
        for fp in file_paths:
            path = Path(fp)
            if not path.exists():
                continue
            # 显式使用 openpyxl 引擎读取所有Sheet
            excel_file = pd.ExcelFile(path, engine='openpyxl')
            for sheet_name in excel_file.sheet_names:
                key = f"{path.stem}_{sheet_name}"
                # parse 不改变原始列名和行顺序，保留各列原始类型
                df = excel_file.parse(sheet_name)
                result[key] = df
                self.dataframes[key] = df
        return result

    def read_all_sheets_raw(self, file_paths: List[str]) -> Dict[str, Dict[str, Any]]:
        """
        返回每个Sheet的原始结构信息，不修改原始数据。

        使用 openpyxl 直接读取工作簿元数据，避免 pandas 类型推断影响。
        每条记录包含：
          - sheet_name: Sheet名称
          - rows: 总行数（不含表头）
          - columns: 总列数
          - column_names: 列名列表
          - column_dtypes: {列名: openpyxl单元格类型描述}
          - sample_data: 前3行数据（原始值列表）

        参数:
            file_paths: Excel文件路径列表
        返回:
            {文件名_Sheet名: 结构信息字典} 字典
        """
        result = {}
        for fp in file_paths:
            path = Path(fp)
            if not path.exists():
                continue
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                key = f"{path.stem}_{sheet_name}"

                # 将工作表数据转为列表（每行一个列表）
                rows_data = list(ws.iter_rows(values_only=True))
                if not rows_data:
                    result[key] = {
                        "sheet_name": sheet_name,
                        "rows": 0,
                        "columns": 0,
                        "column_names": [],
                        "column_dtypes": {},
                        "sample_data": []
                    }
                    continue

                # 第一行为表头
                headers = [str(cell) if cell is not None else f"Unnamed:{i}" for i, cell in enumerate(rows_data[0])]
                num_cols = len(headers)

                # 按列收集数据类型（从数据行推断）
                column_dtypes = {}
                for col_idx in range(num_cols):
                    col_name = headers[col_idx]
                    col_types = set()
                    for row in rows_data[1:]:
                        if col_idx < len(row) and row[col_idx] is not None:
                            col_types.add(type(row[col_idx]).__name__)
                    column_dtypes[col_name] = sorted(col_types) if col_types else ["NoneType"]

                # 前3行样本（从第2行开始，每行转为 {列名: 值} 格式）
                data_rows = rows_data[1:4]  # 最多取3行
                sample_data = []
                for row in data_rows:
                    row_dict = {}
                    for col_idx, header in enumerate(headers):
                        val = row[col_idx] if col_idx < len(row) else None
                        # 处理无法JSON序列化的类型
                        if isinstance(val, (pd.Timestamp,)):
                            val = str(val)
                        row_dict[header] = val
                    sample_data.append(row_dict)

                result[key] = {
                    "sheet_name": sheet_name,
                    "rows": max(0, len(rows_data) - 1),  # 去掉表头
                    "columns": num_cols,
                    "column_names": headers,
                    "column_dtypes": column_dtypes,
                    "sample_data": sample_data
                }
            wb.close()
        return result

    def clean_data(
        self,
        df: pd.DataFrame,
        fill_strategy: Optional[Dict[str, Any]] = None,
        return_info: bool = False
    ) -> Union[pd.DataFrame, Tuple[pd.DataFrame, Dict[str, Any]]]:
        """
        数据清洗：去重 + 可选的自定义空值填充。

        默认行为：只做去重，不做任何空值填充。
        调用者通过 fill_strategy 自行决定填充策略。

        参数:
            df: 输入DataFrame
            fill_strategy: 可选的自定义填充策略字典。
                           键为 dtype 名称（如 'int64', 'float64', 'object'）
                           或具体列名；值为要填充的值。
                           例如: {"int64": 0, "object": "未知"}
                           为 None 时只去重不填充。
            return_info: 为 True 时返回 (清洗后df, 对比信息字典)

        返回:
            return_info=False: 仅返回清洗后的 DataFrame
            return_info=True: 返回 (清洗后df, 对比信息)
                              对比信息包含:
                              - before_rows: 清洗前行数
                              - after_rows: 清洗后行数
                              - duplicates_removed: 去重行数
                              - filled_cells: 填充的单元格数（按列统计）
        """
        before_rows = len(df)
        before_columns = len(df.columns)
        df_clean = df.copy()

        # 记录空值填充情况
        filled_cells = {}

        # 如果指定了填充策略，则按策略填充
        if fill_strategy is not None:
            for key, value in fill_strategy.items():
                # 按 dtype 填充
                dtype_cols = [c for c in df_clean.columns if df_clean[c].dtype.name == key]
                if dtype_cols:
                    for col in dtype_cols:
                        na_count = int(df_clean[col].isna().sum())
                        if na_count > 0:
                            df_clean[col] = df_clean[col].fillna(value)
                            filled_cells[col] = na_count
                # 按列名精确匹配（覆盖 dtype 策略）
                if key in df_clean.columns:
                    na_count = int(df_clean[key].isna().sum())
                    if na_count > 0:
                        df_clean[key] = df_clean[key].fillna(value)
                        filled_cells[key] = na_count

        # 去重
        df_clean = df_clean.drop_duplicates()
        after_rows = len(df_clean)

        if not return_info:
            return df_clean

        info = {
            "before_rows": before_rows,
            "after_rows": after_rows,
            "duplicates_removed": before_rows - after_rows,
            "filled_cells": filled_cells if filled_cells else None,
            "total_filled": sum(filled_cells.values()) if filled_cells else 0
        }
        return df_clean, info

    def identify_dimensions(self, df: pd.DataFrame) -> Dict[str, List[str]]:
        """
        自动识别数据维度。

        识别策略（优先级从高到低）：
          1. datetime64 类型 → "时间" 维度
          2. 数值类型 (int64/float64) → "数值" 维度
          3. object 类型 → 先通过关键词匹配检查是否为时间相关，否则归为"分类"
          4. 关键词辅助匹配覆盖其他特殊情况

        关键词列表覆盖常见的行业通用词，但不再作为唯一判断依据。

        返回:
            {"时间": [字段列表], "分类": [字段列表], "数值": [字段列表]}
        """
        dimensions = {"时间": [], "分类": [], "数值": []}

        # 行业通用的时间关键词
        date_keywords = [
            "日期", "时间", "年", "月", "日", "时", "季度", "周", "星期",
            "date", "time", "year", "month", "day", "week", "quarter",
            "年度", "月份", "日期时间", "datetime", "timestamp",
            "period", "期间", "开始", "结束", "start", "end", "创建时间",
            "更新时间", "modified", "created", "审批日期", "单据日期"
        ]
        # 行业通用的分类关键词
        category_keywords = [
            "部门", "产品", "区域", "类型", "类别", "渠道", "城市", "品牌",
            "名称", "分类", "状态", "级别", "性别", "色系", "款式", "系列",
            "供应商", "客户", "门店", "仓库", "业务", "行业", "标签",
            "dept", "product", "region", "city", "category", "status",
            "level", "brand", "name", "supplier", "customer", "store",
            "warehouse", "segment", "group", "industry", "business", "tag",
            "省份", "地址", "地区", "单位", "规格", "型号", "备注", "说明",
            "province", "address", "area", "unit", "spec", "model", "remark",
            "负责人", "经理", "owner", "manager", "员工", "employee"
        ]

        for col in df.columns:
            col_lower = str(col).lower()
            col_dtype = df[col].dtype

            # 步骤1: datetime 类型 → 时间
            if pd.api.types.is_datetime64_any_dtype(col_dtype):
                dimensions["时间"].append(col)
                continue

            # 步骤2: 数值类型 → 数值
            if pd.api.types.is_numeric_dtype(col_dtype):
                dimensions["数值"].append(col)
                continue

            # 步骤3: object（文本）类型 → 先用关键词辅助判断
            if pd.api.types.is_object_dtype(col_dtype):
                # 检查是否为"时间"相关（年月日等关键词可能出现在文本列中）
                if any(kw in col_lower for kw in date_keywords):
                    dimensions["时间"].append(col)
                    continue
                # 检查是否为"分类"相关
                if any(kw in col_lower for kw in category_keywords):
                    dimensions["分类"].append(col)
                    continue
                # 无匹配关键词的 object 列，默认归为分类
                dimensions["分类"].append(col)
                continue

            # 步骤4: 其他类型（bool, timedelta 等）归为分类
            dimensions["分类"].append(col)

        return dimensions

    def aggregate_metrics(
        self,
        df: pd.DataFrame,
        dims: Dict[str, List[str]]
    ) -> Dict[str, Any]:
        """
        按维度聚合计算指标。

        参数:
            df: 数据源
            dims: 维度字典（来自 identify_dimensions 的输出）

        返回:
            聚合结果字典，包含：
            - total: 各数值列的统计（sum/avg/max/min/count）
            - by_dimension: 按各分类维度分组聚合的结果
            - insights: 自动生成的数据结论列表
        """
        result: Dict[str, Any] = {
            "total": {},
            "by_dimension": {},
            "insights": []
        }

        # 计算数值列的合计和均值
        for col in dims.get("数值", []):
            if col in df.columns:
                series = df[col]
                result["total"][col] = {
                    "sum": float(series.sum()) if not series.isna().all() else 0.0,
                    "avg": float(series.mean()) if not series.isna().all() else 0.0,
                    "max": float(series.max()) if not series.isna().all() else 0.0,
                    "min": float(series.min()) if not series.isna().all() else 0.0,
                    "count": int(series.count())
                }

        # 按分类维度分组聚合
        numeric_cols = dims.get("数值", [])
        for dim in dims.get("分类", []):
            if dim in df.columns and numeric_cols:
                # 只对确实存在于 df 中的数值列分组
                valid_metrics = [c for c in numeric_cols if c in df.columns]
                if valid_metrics:
                    grouped = df.groupby(dim)[valid_metrics].sum()
                    result["by_dimension"][dim] = grouped.to_dict()

        # 自动生成数据结论
        if dims.get("数值"):
            top_metric = dims["数值"][0]
            if top_metric in result["total"]:
                total_val = result["total"][top_metric]["sum"]
                count_val = result["total"][top_metric]["count"]
                result["insights"].append(f"{top_metric}总计: {total_val:,.2f}")
                result["insights"].append(f"{top_metric}有效记录数: {count_val}")

        return result

    def calculate_comparison(self, current: float, previous: float) -> Dict[str, float]:
        """
        计算同比/环比增长率。

        参数:
            current: 当前期数值
            previous: 基期数值

        返回:
            {"yoy": 同比增长率(%), "qoq": 环比增长率(%)}
        """
        yoy = ((current - previous) / previous * 100) if previous != 0 else 0.0
        qoq = ((current - previous) / previous * 100) if previous != 0 else 0.0
        return {"yoy": round(yoy, 2), "qoq": round(qoq, 2)}
